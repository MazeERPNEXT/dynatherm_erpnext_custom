        
import frappe
import math
from openpyxl import load_workbook
from frappe.utils import flt
from erpnext.manufacturing.doctype.bom_creator.bom_creator import BOMCreator


# =========================================================
# COMMON CALCULATION FUNCTION (FROM bom.js LOGIC)
# =========================================================
@frappe.whitelist()
def recalc_item(item):

    import json

    if isinstance(item, str):
        item = json.loads(item)

    return calculate_values(item)

def calculate_values(item):

    import math
    from frappe.utils import flt

    def get(key):
        if isinstance(item, dict):
            return flt(item.get(key))
        return flt(getattr(item, key, 0))

    def setv(key, value):
        if isinstance(item, dict):
            item[key] = value
        else:
            setattr(item, key, value)

    density = flt(get("custom_density"))
    qty = flt(get("qty"))

    item_group = (
        item.get("item_group") if isinstance(item, dict)
        else getattr(item, "item_group", None)
    )

    shape = (
        item.get("custom_shape") if isinstance(item, dict)
        else getattr(item, "custom_shape", None)
    )

    L = flt(get("custom_length"))
    W = flt(get("custom_width"))
    T = flt(get("custom_thickness"))
    OD = flt(get("custom_outer_diameter"))
    ID = flt(get("custom_inner_diameter"))
    WALL = flt(get("custom_wall_thickness"))

    π = math.pi
    base = 0

    if not density:
        setv("custom_kilogramskgs", 0)
        setv("custom_total_weight", 0)
        return item

    # ================= PLATES =================
    if item_group == "Plates":

        if shape == "Rectangle" and L and W and T:
            base = (L * W * T * density) / 1_000_000

        elif shape == "Circle" and OD and T:
            base = (π * (OD / 2) ** 2 * T * density) / 1_000_000

        elif shape == "Hollow":
            if ID and T and L:

                OD_calc = OD if OD else (ID + 2 * T)
                outer = (OD_calc / 2) ** 2
                inner = (ID / 2) ** 2
                base = (π * (outer - inner) * L * density) / 1_000_000

    # ================= PIPES / TUBES =================
    # elif item_group in ["Pipes", "Tubes"]:

    #     if OD and WALL and L:
    #         R = OD / 2
    #         r = max(R - WALL, 0)
    #         base = (π * (R**2 - r**2) * L * density) / 1_000_000
    elif item_group in ["Pipes", "Tubes"]:

        if shape == "Hollow" and OD and T and L:
            ID_calc = OD - (2 * T)

            if ID_calc > 0:
                base = (π * ((OD / 2) ** 2 - (ID_calc / 2) ** 2) * L * density) / 1_000_000

    # ================= FORGINGS =================
    elif item_group == "Forgings":

        if shape == "Hollow" and OD and T and L:
            ID_calc = OD - (2 * T)
            base = (π * ((OD/2)**2 - (ID_calc/2)**2) * L * density) / 1_000_000

        elif shape == "Circle" and OD and T:
            base = (π * (OD / 2) ** 2 * T * density) / 1_000_000

    # ================= RODS =================
    elif item_group == "Rods":

        if shape == "Circle" and OD and L:
            base = (π * (OD / 2) ** 2 * L * density) / 1_000_000

    # ================= FLANGES =================
    elif item_group in ["Flanges", "Rings"]:

        if OD and ID and T:
            base = (π * ((OD/2)**2 - (ID/2)**2) * T * density) / 1_000_000

    # ================= FINAL =================
    kg_per_unit = flt(base, 4)
    total = flt(qty * kg_per_unit, 4)

    setv("custom_kilogramskgs", kg_per_unit)
    setv("custom_total_weight", total)

    scrap_pct = flt(get("custom_scrap_margin_percentage"))
    transport_rate = flt(get("custom_transportation_cost"))

    setv("custom_scrap_margin_kgs", flt(total * scrap_pct / 100, 4))
    setv("custom_transportation_cost_kgs", flt(total * transport_rate, 2))

    return item

# =========================================================
# EXCEL UPLOAD METHOD
# =========================================================
@frappe.whitelist()
def upload_bom_excel(file_url):

    if not file_url:
        frappe.throw("File URL is missing")

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_path = file_doc.get_full_path()

    wb = load_workbook(file_path)
    sheet = wb.active

    data = []
    headers = [cell.value for cell in sheet[1]]

    field_map = {
        "Item Code": "item_code",
        "Item Name": "item_name",
        "Item Group": "item_group",
        "Shape": "custom_shape",
        "Finished Goods Item": "fg_item",
        "Is Expandable": "is_expandable",
        "BOM Created": "bom_created",
        "Qty": "qty",
        "Rate": "rate",
        "UOM": "uom",
        "Parent Row No": "parent_row_no",

        "Part Number": "custom_part_number",
        "Length (mm)": "custom_length",
        "Width (mm)": "custom_width",
        "Thickness (mm)": "custom_thickness",
        "Density (kg/m³)": "custom_density",
        "Outer Diameter (mm)": "custom_outer_diameter",
        "Inner Diameter (mm)": "custom_inner_diameter",
        "Wall Thickness (mm)": "custom_wall_thickness",
        "Kgs Per Unit": "custom_kilogramskgs",
        "Total Weight": "custom_total_weight",
        "Scrap Margin (%)": "custom_scrap_margin_percentage",
        "Scrap Margin (Kg)": "custom_scrap_margin_kgs",
        "Transportation Cost (₹ / Kg)": "custom_transportation_cost",
        "Transportation Cost (₹)":"custom_transportation_cost_kgs",
        
    }

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))

        if not row_dict.get("Item Code"):
            continue

        child = {}

        for excel_field, frappe_field in field_map.items():
            value = row_dict.get(excel_field)

            if frappe_field in ["qty", "rate"]:
                value = value or 0

            if frappe_field in ["is_expandable", "bom_created"]:
                value = int(value) if value else 0

            if frappe_field.startswith("custom_") and value not in [None, ""]:
                try:
                    value = float(value)
                except:
                    pass

            child[frappe_field] = value

        # AUTO CALCULATION HERE
        child = calculate_values(child)
        data.append(child)
    return data

# =========================================================
# CUSTOM BOM CREATOR
# =========================================================

class CustomBOMCreator(BOMCreator):
    @frappe.whitelist()
    def process_item_selection(self, item_idx=None):
        pass
    
    def validate_duplicate_item(self):
        pass
    
    def validate(self):
        super().validate()

        # your calculations
        for item in self.items:
            calculate_values(item)

    def create_bom(self, row, production_item_wise_rm):

        bom_creator_item = row.name if row.name != self.name else ""

        if frappe.db.exists(
            "BOM",
            {
                "bom_creator": self.name,
                "item": row.item_code,
                "bom_creator_item": bom_creator_item,
                "docstatus": 1,
            },
        ):
            return

        bom = frappe.new_doc("BOM")

        bom.update({
            "item": row.item_code,
            "bom_type": "Production",
            "quantity": row.qty,
            "allow_alternative_item": 1,
            "bom_creator": self.name,
            "bom_creator_item": bom_creator_item,
            "is_phantom_bom": row.get("is_phantom_item"),
        })

        if row.item_code == self.item_code and (self.routing or self.has_operations()):
            bom.routing = self.routing
            bom.with_operations = 1
            bom.transfer_material_against = "Work Order"

        BOM_FIELDS = [
            "company",
            "rm_cost_as_per",
            "project",
            "currency",
            "conversion_rate",
            "buying_price_list",
        ]

        for field in BOM_FIELDS:
            if self.get(field):
                bom.set(field, self.get(field))

        for item in production_item_wise_rm[(row.item_code, row.name)]["items"]:

            # AUTO CALCULATION AGAIN (SAFETY)
            calculate_values(item)

            bom_no = ""
            item.do_not_explode = 1

            if (item.item_code, item.name) in production_item_wise_rm:
                bom_no = production_item_wise_rm.get(
                    (item.item_code, item.name)
                ).bom_no
                item.do_not_explode = 0

            item_args = {
                "item_code": item.item_code,
                "qty": item.qty,
                "uom": item.uom,
                "rate": item.rate,
                "stock_qty": item.stock_qty,
                "stock_uom": item.stock_uom,
                "conversion_factor": item.conversion_factor,
                "do_not_explode": item.do_not_explode,
                "operation": item.operation,
                "is_phantom_item": item.is_phantom_item,
            }

            CUSTOM_FIELDS = [
                "custom_part_number",
                "custom_shape",
                "custom_length",
                "custom_width",
                "custom_thickness",
                "custom_density",
                "custom_outer_diameter",
                "custom_inner_diameter",
                "custom_wall_thickness",
                "custom_kilogramskgs",
                "custom_total_weight",
                "custom_scrap_margin_percentage",
                "custom_scrap_margin_kgs",
                "custom_transportation_cost",
                "custom_transportation_cost_kgs",
            ]

            for field in CUSTOM_FIELDS:
                item_args[field] = item.get(field)

            item_args.update({
                "bom_no": bom_no,
                "allow_alternative_item": 1,
                "allow_scrap_items": not item.get("is_phantom_item"),
                "include_item_in_manufacturing": 1,
            })

            bom.append("items", item_args)

        bom.save(ignore_permissions=True)
        bom.submit()

        production_item_wise_rm[(row.item_code, row.name)].bom_no = bom.name





# import frappe
from openpyxl import Workbook
import io

@frappe.whitelist()
def download_bom_template():

    wb = Workbook()
    ws = wb.active
    ws.title = "BOM Template"

    # =====================================================
    # ✅ HEADERS (MATCH YOUR SYSTEM)
    # =====================================================

    headers = [
        "Part Number",
        "Item Code",
        "Item Group",
        "Shape",
        "Qty",
        "Finished Goods Item",

        # Custom Fields
        "Length (mm)",
        "Width (mm)",
        "Thickness (mm)",
        "Density (kg/m³)",
        "Outer Diameter (mm)",
        "Inner Diameter (mm)",
    ]

    ws.append(headers)

    # =====================================================
    # ✅ AUTO WIDTH
    # =====================================================

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 3

    # =====================================================
    # ✅ SAVE FILE
    # =====================================================

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": "BOM_Creator_Template.xlsx",
        "is_private": 0,
        "content": file_stream.read()
    })

    file_doc.insert(ignore_permissions=True)

    return file_doc.file_url