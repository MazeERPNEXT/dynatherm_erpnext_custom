# Copyright (c) 2026, maze and contributors
# For license information, please see license.txt

# import frappe
# from frappe.model.document import Document


# class CuttingPlan(Document):
# 	pass

import frappe
from frappe.model.document import Document
from openpyxl import load_workbook
from frappe.utils import flt
import math
    
class CuttingPlan(Document):
    # def validate(self):
    #     self.sync_job_no_from_child()
    
    def after_insert(self):
        if not self.cutting_no:
            self.db_set("cutting_no", self.name)

    def sync_job_no_from_child(self):
        job_nos = set()

        # 🔹 Collect from child table
        for row in self.cutting_plan_plate_details:
            if row.job_no:
                job_nos.add(row.job_no)

        # 🔹 Clear parent table
        self.set("job_no", [])

        # 🔹 Add unique values to parent multiselect table
        for jn in job_nos:
            self.append("job_no", {
                "job_no": jn
            })
            
@frappe.whitelist()
def make_material_request(cutting_plan):

    cp = frappe.get_doc("Cutting Plan", cutting_plan)

    if cp.docstatus != 1:
        frappe.throw("Cutting Plan must be Submitted")

    # 🔹 Check existing Material Request using Cutting Plan number
    existing_mr = frappe.db.get_value(
        "Material Request Item",
        {"custom_cutting_plan_no": cp.name},
        "parent"
    )

    if existing_mr:
        return existing_mr

    # 🔹 Create new Material Request
    mr = frappe.new_doc("Material Request")
    mr.material_request_type = "Purchase"
    mr.schedule_date = cp.date

    # Loop Cutting Plan Plate Details
    for row in cp.cutting_plan_plate_details:
         # ✅ Only send items where material is To Be Ordered
        if row.material_availability != "To Be Ordered":
            continue
        
        mr.append("items", {
            "custom_cutting_plan_no": cp.name,
            "item_code": row.item_code,
            "qty": row.qty or 1,
            "uom": row.uom,
            "schedule_date": cp.date,
            "custom_length": row.length,
            "custom_width": row.width,
            "custom_thickness": row.thickness,
            "custom_density": row.density,
            "custom_kilogramskgs": row.kgs_per_unit,
            "custom_total_weight": row.total_weight,
            
        })

    mr.insert(ignore_permissions=True)
    return mr.name


@frappe.whitelist()
def make_request_for_quotation(cutting_plan):

    cp = frappe.get_doc("Cutting Plan", cutting_plan)

    if cp.docstatus != 1:
        frappe.throw("Cutting Plan must be Submitted")

    rfq = frappe.new_doc("Request for Quotation")

    rfq.transaction_date = cp.date
    rfq.company = frappe.defaults.get_user_default("Company")
    rfq.status = "Draft"

    for row in cp.cutting_plan_plate_details:

        if row.material_availability != "To Be Ordered":
            continue

        rfq.append("items", {
            "custom_cutting_plan_no": cp.name,
            "item_code": row.item_code,
            "qty": row.qty or 1,
            "uom": row.uom,
            "conversion_factor": 1,
            "custom_length": row.length,
            "custom_width": row.width,
            "custom_thickness": row.thickness,
        })

    # 🔥 REQUIRED FIXES
    rfq.flags.ignore_validate = True
    rfq.flags.ignore_mandatory = True   

    rfq.insert(ignore_permissions=True)
    return rfq.name



@frappe.whitelist()
def get_cutting_plans(project):
    data = frappe.db.sql("""
        SELECT 
            cp.name AS cutting_plan_no,
            cp.date
        FROM `tabCutting Plan` cp
        INNER JOIN `tabCutting Plan Item` cpi
            ON cpi.parent = cp.name
        WHERE cpi.project = %s
    """, (project,), as_dict=1)

    return data


@frappe.whitelist()
def recalc_item(item):
    import json

    if isinstance(item, str):
        item = json.loads(item)

    return calculate_cutting_plan_values(item)

def calculate_cutting_plan_values(item):

    def get(key):
        if isinstance(item, dict):
            return flt(item.get(key))
        return flt(getattr(item, key, 0))

    def setv(key, value):
        if isinstance(item, dict):
            item[key] = value
        else:
            setattr(item, key, value)

    density = get("density")
    qty = get("qty")

    item_group = item.get("item_group") if isinstance(item, dict) else getattr(item, "item_group", None)
    shape = item.get("shape") if isinstance(item, dict) else getattr(item, "shape", None)

    L = get("length")
    W = get("width")
    T = get("thickness")
    OD = get("outer_diameter")
    ID = get("inner_diameter")

    π = math.pi
    base = 0

    if not density:
        setv("kgs_per_unit", 0)
        setv("total_weight", 0)
        return item

    # ================= PLATES =================
    if item_group == "Plates":

        if shape == "Rectangle" and L and W and T:
            base = (L * W * T * density) / 1_000_000

        elif shape == "Circle" and OD and T:
            base = (π * (OD / 2) ** 2 * T * density) / 1_000_000

        elif shape == "Hollow" and ID and T and L:
            OD_calc = ID + (2 * T)
            base = (π * ((OD_calc/2)**2 - (ID/2)**2) * L * density) / 1_000_000

    # ================= PIPES / TUBES =================
    elif item_group in ["Pipes", "Tubes"]:

        if shape == "Hollow" and OD and T and L:
            ID_calc = OD - (2 * T)
            if ID_calc > 0:
                base = (π * ((OD/2)**2 - (ID_calc/2)**2) * L * density) / 1_000_000

    # ================= FINAL =================
    kg_per_unit = flt(base, 4)
    total = flt(qty * kg_per_unit, 4)

    setv("kgs_per_unit", kg_per_unit)
    setv("total_weight", total)

    return item

@frappe.whitelist()
def upload_cutting_plan_excel(file_url):

    if not file_url:
        frappe.throw("File URL is missing")

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_path = file_doc.get_full_path()

    wb = load_workbook(file_path)
    sheet = wb.active

    data = []
    headers = [cell.value for cell in sheet[1]]

    # ✅ MAP YOUR EXCEL COLUMNS → DOCTYPE FIELDS
    field_map = {
        "Plate Number": "plate_number",
        "Item Code": "item_code",
        "Item Name": "item_name",
        "Item Group": "item_group",
        "Shape": "shape",
        "Tag No": "tag_no",
        "Qty": "qty",
        # "UOM": "uom",
        "Job No": "job_no",
        "Material Availability": "material_availability",

        "Length (mm)": "length",
        "Width (mm)": "width",
        "Thickness (mm)": "thickness",
        "Outer Diameter (mm)": "outer_diameter",
        "Inner Diameter (mm)": "inner_diameter",
        "Density (kg/m³)": "density",
    }

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))

        if not row_dict.get("Item Code"):
            continue

        child = {}

        for excel_field, frappe_field in field_map.items():
            value = row_dict.get(excel_field)

            if frappe_field in ["qty"]:
                value = flt(value or 0)

            child[frappe_field] = value
        child = calculate_cutting_plan_values(child)

        data.append(child)

    return data

@frappe.whitelist()
def get_bom_from_project(project):

    bom = frappe.db.get_value(
        "BOM",
        {
            "project": project,
            "docstatus": 1   # only submitted BOM
        },
        ["name", "item"],
        as_dict=True
    )

    if not bom:
        return {}

    return {
        "tag_no": bom.item,
        "bom_no": bom.name
    }
    
    
    
@frappe.whitelist()
def get_part_numbers_from_bom(bom_no):

    if not bom_no:
        return []

    items = frappe.get_all(
        "BOM Item",
        filters={"parent": bom_no},
        fields=["custom_part_number", "idx"],  # ✅ include idx
        order_by="idx asc"                     # ✅ maintain order
    )

    part_numbers = []
    seen = set()

    for d in items:
        pn = d.get("custom_part_number")
        if pn and pn not in seen:
            part_numbers.append(pn)
            seen.add(pn)

    return part_numbers


@frappe.whitelist()
def get_part_details(bom_no, part_number):

    if not bom_no or not part_number:
        return {}

    # ===============================
    # FETCH ALL ITEMS FIRST
    # ===============================
    items = frappe.get_all(
        "BOM Item",
        filters={
            "parent": bom_no
        },
        fields=[
            "item_code",
            "qty",
            "uom",
            "custom_item_group",
            "custom_shape",
            "custom_length",
            "custom_width",
            "custom_thickness",
            "custom_density",
            "custom_outer_diameter",
            "custom_inner_diameter",
            "custom_kilogramskgs",
            "custom_total_weight",
            "custom_part_number"
        ]
    )

    matched = None

    # ===============================
    # SAFE STRING MATCH
    # ===============================
    for d in items:

        if str(d.custom_part_number).strip() == str(part_number).strip():
            matched = d
            break

    if not matched:
        return {}

    return {
        "item_code": matched.item_code,
        "qty": matched.qty,
        "uom": matched.uom,
        "item_group": matched.custom_item_group,
        "shape": matched.custom_shape,
        "length": matched.custom_length,
        "width": matched.custom_width,
        "thickness": matched.custom_thickness,
        "density": matched.custom_density,
        "outer_diameter": matched.custom_outer_diameter,
        "inner_diameter": matched.custom_inner_diameter,
        "kgs_per_unit": matched.custom_kilogramskgs,
        "total_weight": matched.custom_total_weight
    }
    
# @frappe.whitelist()
# def get_part_details(bom_no, part_number):

#     if not bom_no or not part_number:
#         return {}

#     item = frappe.get_value(
#         "BOM Item",
#         {
#             "parent": bom_no,
#             "custom_part_number": part_number
#         },
#         ["item_code", "qty", "uom", "custom_item_group", "custom_shape", "custom_length", "custom_width", "custom_thickness", "custom_density", "custom_outer_diameter", "custom_inner_diameter", "custom_kilogramskgs", "custom_total_weight"],
#         as_dict=True
#     )

#     if not item:
#         return {}

#     return {
#         "item_code":item.item_code,
#         "qty":item.qty,
#         "uom":item.uom,
#         "item_group":item.custom_item_group,
#         "shape":item.custom_shape,
#         "length":item.custom_length,
#         "width":item.custom_width,
#         "thickness":item.custom_thickness,
#         "density":item.custom_density,
#         "outer_diameter":item.custom_outer_diameter,
#         "inner_diameter":item.custom_inner_diameter,
#         "kgs_per_unit": item.custom_kilogramskgs,
#         "total_weight": item.custom_total_weight
#     }